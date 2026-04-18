"""
Tomorrow World – Online Yacht Booking System
"""
import io
import os
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, send_file)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from flask_mail import Mail, Message

from config import Config
from models import db, User, Yacht, Booking, Expense

def _inject_now():
    return dict(now=datetime.now())

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
mail = Mail(app)

login_manager = LoginManager(app)
login_manager.login_view = 'auth_login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'warning'

app.context_processor(_inject_now)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ---------------------------------------------------------------------------
# Access decorators
# ---------------------------------------------------------------------------
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


def staff_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_staff:
            flash('This page is for staff only. Please log in.', 'danger')
            return redirect(url_for('auth_login'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Database seed
# ---------------------------------------------------------------------------
def init_db():
    db.create_all()

    if not User.query.filter_by(role='admin').first():
        admin = User(username='admin', email='admin@tomorrowworld.com',
                     name='System Admin', role='admin')
        admin.set_password(app.config.get('INIT_ADMIN_PASSWORD', 'Admin@123'))
        db.session.add(admin)

    if not User.query.filter_by(username='staff1').first():
        s = User(username='staff1', email='staff1@tomorrowworld.com',
                 name='David Chan', department='Sales & Marketing',
                 job_title='Sales Executive', role='staff')
        s.set_password('Staff@123')
        db.session.add(s)

    if not Yacht.query.first():
        yachts_data = [
            dict(
                name='TW Horizon I',
                model='Sunseeker Predator 74',
                capacity=20,
                length_m=22.5,
                description='Our flagship luxury yacht, fully equipped with premium entertainment facilities. Perfect for corporate hospitality, product launches, and private parties.',
                price_per_hour=5000,
                image_url='https://images.unsplash.com/photo-1567899378494-47b22a2ae96a?auto=format&fit=crop&w=800&q=80',
            ),
            dict(
                name='TW Horizon II',
                model='Princess V78',
                capacity=15,
                length_m=23.8,
                description='Mid-size luxury yacht ideal for intimate business meetings, team building events, and family gatherings. Equipped with a full galley and spacious deck.',
                price_per_hour=4000,
                image_url='https://images.unsplash.com/photo-1544551763-46a013bb70d5?auto=format&fit=crop&w=800&q=80',
            ),
            dict(
                name='TW Pearl',
                model='Azimut 50',
                capacity=10,
                length_m=15.2,
                description='Our elegant compact yacht, perfect for sunset cruises, romantic dinners, and small private gatherings. Agile and stylish.',
                price_per_hour=2500,
                image_url='https://images.unsplash.com/photo-1605281317010-fe5ffe798166?auto=format&fit=crop&w=800&q=80',
            ),
        ]
        for d in yachts_data:
            db.session.add(Yacht(**d))

    db.session.commit()


# ---------------------------------------------------------------------------
# Public routes
# ---------------------------------------------------------------------------
@app.route('/')
def index():
    yachts = Yacht.query.filter_by(is_available=True).all()
    total_bookings = Booking.query.filter_by(status='approved').count()
    return render_template('index.html', yachts=yachts, total_bookings=total_bookings)


@app.route('/booking/customer', methods=['GET', 'POST'])
def booking_customer():
    yachts = Yacht.query.filter_by(is_available=True).all()
    if request.method == 'POST':
        try:
            booking = Booking(
                booking_ref=Booking.generate_ref(),
                booking_type='customer',
                project=request.form.get('project', '').strip(),
                unit_number=request.form.get('unit_number', '').strip(),
                customer_name=request.form['customer_name'].strip(),
                customer_email=request.form['customer_email'].strip().lower(),
                customer_phone=request.form.get('customer_phone', '').strip(),
                client_company=request.form.get('client_company', '').strip(),
                yacht_id=int(request.form['yacht_id']) if request.form.get('yacht_id') else None,
                booking_date=datetime.strptime(request.form['booking_date'], '%Y-%m-%d').date(),
                start_time=datetime.strptime(request.form['start_time'], '%H:%M').time(),
                end_time=datetime.strptime(request.form['end_time'], '%H:%M').time(),
                num_passengers=int(request.form.get('num_passengers', 1)),
                destination=request.form.get('destination', '').strip(),
                has_alcohol='has_alcohol' in request.form,
                has_pizza='has_pizza' in request.form,
                has_vegetarian='has_vegetarian' in request.form,
                is_offshore='is_offshore' in request.form,
                food_allergies=request.form.get('food_allergies', '').strip(),
                marketing_support='marketing_support' in request.form,
                other_requests=request.form.get('other_requests', '').strip(),
            )
            db.session.add(booking)
            db.session.commit()
            _send_confirmation(booking)
            _send_admin_notification(booking)
            flash(f'Booking submitted successfully! Your reference: {booking.booking_ref}', 'success')
            return redirect(url_for('booking_success', ref=booking.booking_ref))
        except Exception as e:
            db.session.rollback()
            import traceback
            app.logger.error('Customer booking submission failed:\n' + traceback.format_exc())
            flash(f'Submission failed: {e}', 'danger')
    return render_template('booking_customer.html', yachts=yachts)


@app.route('/booking/staff', methods=['GET', 'POST'])
@login_required
@staff_required
def booking_staff():
    yachts = Yacht.query.filter_by(is_available=True).all()
    if request.method == 'POST':
        try:
            booking = Booking(
                booking_ref=Booking.generate_ref(),
                booking_type='internal',
                project=request.form.get('project', '').strip(),
                unit_number=request.form.get('unit_number', '').strip(),
                user_id=current_user.id,
                customer_name=current_user.name or request.form['applicant_name'].strip(),
                customer_email=current_user.email,
                customer_phone=request.form.get('customer_phone', '').strip(),
                client_company=request.form.get('client_company', '').strip(),
                department=request.form.get('department', current_user.department or '').strip(),
                job_title=request.form.get('job_title', current_user.job_title or '').strip(),
                supervisor=request.form.get('supervisor', '').strip(),
                yacht_id=int(request.form['yacht_id']) if request.form.get('yacht_id') else None,
                booking_date=datetime.strptime(request.form['booking_date'], '%Y-%m-%d').date(),
                start_time=datetime.strptime(request.form['start_time'], '%H:%M').time(),
                end_time=datetime.strptime(request.form['end_time'], '%H:%M').time(),
                num_passengers=int(request.form.get('num_passengers', 1)),
                destination=request.form.get('destination', '').strip(),
                has_alcohol='has_alcohol' in request.form,
                has_pizza='has_pizza' in request.form,
                has_vegetarian='has_vegetarian' in request.form,
                is_offshore='is_offshore' in request.form,
                food_allergies=request.form.get('food_allergies', '').strip(),
                marketing_support='marketing_support' in request.form,
                other_requests=request.form.get('other_requests', '').strip(),
            )
            db.session.add(booking)
            db.session.commit()
            _send_confirmation(booking)
            _send_admin_notification(booking)
            flash(f'Internal request submitted! Reference: {booking.booking_ref}', 'success')
            return redirect(url_for('booking_success', ref=booking.booking_ref))
        except Exception as e:
            db.session.rollback()
            flash(f'Submission failed: {str(e)}', 'danger')
    return render_template('booking_staff.html', yachts=yachts, user=current_user)


@app.route('/booking/success/<ref>')
def booking_success(ref):
    booking = Booking.query.filter_by(booking_ref=ref).first_or_404()
    return render_template('booking_success.html', booking=booking)


@app.route('/booking/check')
def booking_check():
    ref = request.args.get('ref', '').strip().upper()
    booking = Booking.query.filter_by(booking_ref=ref).first() if ref else None
    return render_template('booking_check.html', booking=booking, ref=ref)


@app.route('/calendar')
def calendar():
    yachts = Yacht.query.filter_by(is_available=True).all()
    return render_template('calendar.html', yachts=yachts)


@app.route('/api/calendar')
def api_calendar():
    yacht_id = request.args.get('yacht_id', '')
    query = Booking.query.filter(Booking.status.in_(['pending', 'approved']))
    if yacht_id:
        query = query.filter_by(yacht_id=int(yacht_id))
    events = []
    for b in query.all():
        color = '#ffc107' if b.status == 'pending' else '#1B7A4A'
        events.append({
            'id': b.id,
            'title': f'{b.yacht.name if b.yacht else "TBC"} | {b.num_passengers} guests',
            'start': f'{b.booking_date}T{b.start_time}',
            'end': f'{b.booking_date}T{b.end_time}',
            'color': color,
            'extendedProps': {'status': b.status, 'ref': b.booking_ref},
        })
    return jsonify(events)


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def auth_login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.is_active and user.check_password(password):
            login_user(user, remember='remember' in request.form)
            flash(f'Welcome back, {user.name or user.username}!', 'success')
            return redirect(request.args.get('next') or url_for('index'))
        flash('Invalid username or password.', 'danger')
    return render_template('auth/login.html')


@app.route('/logout')
@login_required
def auth_logout():
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('index'))


# ---------------------------------------------------------------------------
# Admin
# ---------------------------------------------------------------------------
@app.route('/admin')
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    # Month-to-date expense total (all currencies summed together — assumes AED default)
    today = date.today()
    month_start = today.replace(day=1)
    month_expense_total = db.session.query(
        db.func.coalesce(db.func.sum(Expense.amount), 0)
    ).filter(Expense.expense_date >= month_start).scalar() or 0

    stats = {
        'total':     Booking.query.count(),
        'pending':   Booking.query.filter_by(status='pending').count(),
        'approved':  Booking.query.filter_by(status='approved').count(),
        'rejected':  Booking.query.filter_by(status='rejected').count(),
        'yachts':    Yacht.query.count(),
        'customers': db.session.query(Booking.customer_email).distinct().count(),
        'month_expense_total': float(month_expense_total),
        'month_label': today.strftime('%b %Y'),
    }
    recent = Booking.query.order_by(Booking.created_at.desc()).limit(10).all()
    return render_template('admin/dashboard.html', stats=stats, recent=recent)


@app.route('/admin/bookings')
@login_required
@admin_required
def admin_bookings():
    status_f = request.args.get('status', '')
    type_f   = request.args.get('type', '')
    search   = request.args.get('search', '').strip()
    q = Booking.query
    if status_f:
        q = q.filter_by(status=status_f)
    if type_f:
        q = q.filter_by(booking_type=type_f)
    if search:
        q = q.filter(db.or_(
            Booking.customer_name.ilike(f'%{search}%'),
            Booking.booking_ref.ilike(f'%{search}%'),
            Booking.customer_email.ilike(f'%{search}%'),
            Booking.client_company.ilike(f'%{search}%'),
        ))
    bookings = q.order_by(Booking.created_at.desc()).all()
    return render_template('admin/bookings.html', bookings=bookings,
                           status_f=status_f, type_f=type_f, search=search)


@app.route('/admin/bookings/<int:bid>')
@login_required
@admin_required
def admin_booking_detail(bid):
    booking = Booking.query.get_or_404(bid)
    return render_template(
        'admin/booking_detail.html',
        booking=booking,
        expense_categories=Expense.CATEGORIES,
        today=date.today().isoformat(),
    )


@app.route('/admin/bookings/<int:bid>/approve', methods=['POST'])
@login_required
@admin_required
def admin_approve(bid):
    b = Booking.query.get_or_404(bid)
    b.status = 'approved'
    b.approved_by_id = current_user.id
    b.approved_at = datetime.utcnow()
    b.admin_notes = request.form.get('notes', '').strip()
    db.session.commit()
    _send_status_email(b)
    flash(f'Booking {b.booking_ref} has been approved.', 'success')
    return redirect(url_for('admin_bookings'))


@app.route('/admin/bookings/<int:bid>/reject', methods=['POST'])
@login_required
@admin_required
def admin_reject(bid):
    b = Booking.query.get_or_404(bid)
    b.status = 'rejected'
    b.rejection_reason = request.form.get('reason', '').strip()
    b.admin_notes = request.form.get('notes', '').strip()
    db.session.commit()
    _send_status_email(b)
    flash(f'Booking {b.booking_ref} has been rejected.', 'warning')
    return redirect(url_for('admin_bookings'))


@app.route('/admin/bookings/<int:bid>/cancel', methods=['POST'])
@login_required
@admin_required
def admin_cancel(bid):
    b = Booking.query.get_or_404(bid)
    reason = request.form.get('reason', '').strip()
    b.status = 'cancelled'
    if reason:
        b.admin_notes = (b.admin_notes + '\n' if b.admin_notes else '') + f'Cancelled: {reason}'
    db.session.commit()
    _send_cancellation(b, cancelled_by='admin', reason=reason)
    flash(f'Booking {b.booking_ref} has been cancelled.', 'info')
    return redirect(url_for('admin_bookings'))


# ---------------------------------------------------------------------------
# Expense CRUD (admin-only)
# ---------------------------------------------------------------------------
def _parse_amount(raw):
    """Return a Decimal(12,2) from user input. Raises ValueError on bad input."""
    if raw is None or str(raw).strip() == '':
        raise ValueError('Amount is required.')
    try:
        d = Decimal(str(raw).replace(',', '').strip())
    except (InvalidOperation, ValueError):
        raise ValueError('Amount must be a number.')
    if d < 0:
        raise ValueError('Amount cannot be negative.')
    return d.quantize(Decimal('0.01'))


def _parse_date(raw, fallback=None):
    """YYYY-MM-DD -> date. Falls back if blank."""
    if not raw or not str(raw).strip():
        return fallback or date.today()
    try:
        return datetime.strptime(raw.strip(), '%Y-%m-%d').date()
    except ValueError:
        raise ValueError('Date must be in YYYY-MM-DD format.')


@app.route('/admin/bookings/<int:bid>/expenses/add', methods=['POST'])
@login_required
@admin_required
def admin_expense_add(bid):
    booking = Booking.query.get_or_404(bid)
    try:
        e = Expense(
            booking_id=booking.id,
            category=request.form.get('category', 'Other').strip() or 'Other',
            description=request.form.get('description', '').strip(),
            amount=_parse_amount(request.form.get('amount')),
            currency=request.form.get('currency', 'AED').strip().upper() or 'AED',
            expense_date=_parse_date(request.form.get('expense_date')),
            receipt_ref=request.form.get('receipt_ref', '').strip() or None,
            notes=request.form.get('notes', '').strip() or None,
            recorded_by_id=current_user.id,
        )
        if not e.description:
            raise ValueError('Description is required.')
        if e.category not in Expense.CATEGORIES:
            e.category = 'Other'
        db.session.add(e)
        db.session.commit()
        flash(f'Expense added to booking {booking.booking_ref}.', 'success')
    except ValueError as ex:
        flash(str(ex), 'danger')
    return redirect(url_for('admin_booking_detail', bid=booking.id) + '#expenses')


@app.route('/admin/expenses/<int:eid>/edit', methods=['POST'])
@login_required
@admin_required
def admin_expense_edit(eid):
    e = Expense.query.get_or_404(eid)
    try:
        e.category = request.form.get('category', e.category).strip() or 'Other'
        if e.category not in Expense.CATEGORIES:
            e.category = 'Other'
        new_desc = request.form.get('description', '').strip()
        if not new_desc:
            raise ValueError('Description is required.')
        e.description = new_desc
        e.amount = _parse_amount(request.form.get('amount'))
        e.currency = (request.form.get('currency', e.currency) or 'AED').strip().upper()
        e.expense_date = _parse_date(request.form.get('expense_date'), fallback=e.expense_date)
        e.receipt_ref = request.form.get('receipt_ref', '').strip() or None
        e.notes = request.form.get('notes', '').strip() or None
        db.session.commit()
        flash('Expense updated.', 'success')
    except ValueError as ex:
        flash(str(ex), 'danger')
    return redirect(url_for('admin_booking_detail', bid=e.booking_id) + '#expenses')


@app.route('/admin/expenses/<int:eid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_expense_delete(eid):
    e = Expense.query.get_or_404(eid)
    bid = e.booking_id
    db.session.delete(e)
    db.session.commit()
    flash('Expense deleted.', 'info')
    return redirect(url_for('admin_booking_detail', bid=bid) + '#expenses')


@app.route('/admin/expenses')
@login_required
@admin_required
def admin_expenses():
    """Company-wide expense ledger with filters."""
    category = request.args.get('category', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    search = request.args.get('search', '').strip()

    q = Expense.query.join(Booking, Expense.booking_id == Booking.id)

    if category:
        q = q.filter(Expense.category == category)
    if date_from:
        try:
            q = q.filter(Expense.expense_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            flash('Invalid "from" date; ignored.', 'warning')
    if date_to:
        try:
            q = q.filter(Expense.expense_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            flash('Invalid "to" date; ignored.', 'warning')
    if search:
        q = q.filter(db.or_(
            Expense.description.ilike(f'%{search}%'),
            Expense.receipt_ref.ilike(f'%{search}%'),
            Booking.booking_ref.ilike(f'%{search}%'),
            Booking.customer_name.ilike(f'%{search}%'),
        ))

    expenses = q.order_by(Expense.expense_date.desc(), Expense.created_at.desc()).all()

    # Totals — grouped by currency in case AED ever diverges
    totals = {}
    for e in expenses:
        totals[e.currency] = totals.get(e.currency, Decimal('0')) + (e.amount or Decimal('0'))

    # Per-category breakdown (AED only for the widget — good enough for now)
    category_breakdown = {}
    for e in expenses:
        if e.currency == 'AED':
            category_breakdown[e.category] = category_breakdown.get(e.category, Decimal('0')) + (e.amount or Decimal('0'))

    return render_template(
        'admin/expenses.html',
        expenses=expenses,
        totals=totals,
        category_breakdown=category_breakdown,
        categories=Expense.CATEGORIES,
        filters={
            'category': category, 'date_from': date_from,
            'date_to': date_to, 'search': search,
        },
    )


@app.route('/booking/cancel/<ref>', methods=['POST'])
def booking_cancel(ref):
    """Self-service cancellation from the booking-check page.
    Requires the customer email to match as a simple verification."""
    booking = Booking.query.filter_by(booking_ref=ref.upper()).first_or_404()
    provided_email = request.form.get('email', '').strip().lower()
    if provided_email != booking.customer_email:
        flash('Email does not match our records. Cancellation not processed.', 'danger')
        return redirect(url_for('booking_check', ref=ref))
    if booking.status in ('cancelled', 'rejected'):
        flash('This booking is already closed.', 'warning')
        return redirect(url_for('booking_check', ref=ref))
    reason = request.form.get('reason', '').strip()
    booking.status = 'cancelled'
    if reason:
        booking.admin_notes = (booking.admin_notes + '\n' if booking.admin_notes else '') + \
                              f'Customer-cancelled: {reason}'
    db.session.commit()
    _send_cancellation(booking, cancelled_by='customer', reason=reason)
    flash(f'Your booking {booking.booking_ref} has been cancelled.', 'success')
    return redirect(url_for('booking_check', ref=ref))


@app.route('/admin/bookings/export')
@login_required
@admin_required
def admin_export():
    bookings = Booking.query.order_by(Booking.created_at.desc()).all()
    expenses = Expense.query.order_by(
        Expense.expense_date.desc(), Expense.created_at.desc()
    ).all()

    wb = openpyxl.Workbook()

    # Shared styles
    hdr_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True, name='Arial', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    def _style_header_row(sheet, ncols):
        sheet.row_dimensions[1].height = 30
        for ci in range(1, ncols + 1):
            cell = sheet.cell(row=1, column=ci)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = thin

    def _autosize(sheet):
        for col in sheet.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = \
                min(max_len + 3, 40)
        sheet.freeze_panes = 'A2'

    # ---------- Sheet 1: Booking Records ----------
    ws = wb.active
    ws.title = 'Booking Records'

    headers = [
        'Ref No.', 'Type', 'Customer / Applicant', 'Email', 'Phone', 'Company / Client',
        'Department', 'Job Title', 'Supervisor', 'Yacht', 'Date', 'Start', 'End',
        'Guests', 'Destination', 'Alcohol', 'Pizza', 'Vegetarian', 'Offshore',
        'Food Allergies', 'Marketing Support', 'Other Requests',
        'Status', 'Rejection Reason', 'Admin Notes', 'Submitted', 'Approved At', 'Approved By',
        'Total Expenses (AED)', 'Expense Count'
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    bool_en = lambda v: 'Yes' if v else 'No'
    for b in bookings:
        # Total expenses in AED only (other currencies left out of this aggregate)
        booking_exp_total = sum(
            float(e.amount or 0) for e in b.expenses if (e.currency or 'AED') == 'AED'
        )
        row = [
            b.booking_ref, b.type_label, b.customer_name, b.customer_email,
            b.customer_phone or '', b.client_company or '',
            b.department or '', b.job_title or '', b.supervisor or '',
            b.yacht.name if b.yacht else '',
            str(b.booking_date), str(b.start_time)[:5], str(b.end_time)[:5],
            b.num_passengers or '', b.destination or '',
            bool_en(b.has_alcohol), bool_en(b.has_pizza), bool_en(b.has_vegetarian),
            bool_en(b.is_offshore), b.food_allergies or '', bool_en(b.marketing_support),
            b.other_requests or '', b.status_label,
            b.rejection_reason or '', b.admin_notes or '',
            b.created_at.strftime('%Y-%m-%d %H:%M') if b.created_at else '',
            b.approved_at.strftime('%Y-%m-%d %H:%M') if b.approved_at else '',
            b.approver.name if b.approver else '',
            round(booking_exp_total, 2),
            len(b.expenses),
        ]
        ws.append(row)
        for ci in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=ci).border = thin
    _autosize(ws)

    # ---------- Sheet 2: Expenses ----------
    ws_exp = wb.create_sheet('Expenses')
    exp_headers = [
        'Expense Date', 'Booking Ref', 'Customer', 'Yacht', 'Booking Date',
        'Category', 'Description', 'Amount', 'Currency',
        'Receipt Ref', 'Notes', 'Recorded By', 'Recorded At'
    ]
    ws_exp.append(exp_headers)
    _style_header_row(ws_exp, len(exp_headers))

    for e in expenses:
        b = e.booking
        ws_exp.append([
            str(e.expense_date),
            b.booking_ref if b else '',
            b.customer_name if b else '',
            b.yacht.name if b and b.yacht else '',
            str(b.booking_date) if b else '',
            e.category,
            e.description,
            float(e.amount or 0),
            e.currency or 'AED',
            e.receipt_ref or '',
            e.notes or '',
            e.recorded_by.name if e.recorded_by else '',
            e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
        ])
        for ci in range(1, len(exp_headers) + 1):
            ws_exp.cell(row=ws_exp.max_row, column=ci).border = thin

    # Total row at the bottom (AED subtotal)
    if expenses:
        aed_total = sum(float(e.amount or 0) for e in expenses if (e.currency or 'AED') == 'AED')
        total_row = ['', '', '', '', '', 'TOTAL (AED)', '', round(aed_total, 2),
                     'AED', '', '', '', '']
        ws_exp.append(total_row)
        last_row = ws_exp.max_row
        for ci in range(1, len(exp_headers) + 1):
            c = ws_exp.cell(row=last_row, column=ci)
            c.font = Font(bold=True)
            c.border = thin
    _autosize(ws_exp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'TW_Yacht_Bookings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/admin/yachts', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_yachts():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            y = Yacht(
                name=request.form['name'].strip(),
                model=request.form.get('model', '').strip(),
                capacity=int(request.form.get('capacity', 10)),
                length_m=float(request.form.get('length_m', 0) or 0),
                description=request.form.get('description', '').strip(),
                price_per_hour=float(request.form.get('price_per_hour', 0) or 0),
                image_url=request.form.get('image_url', '').strip(),
                is_available='is_available' in request.form,
            )
            db.session.add(y)
            db.session.commit()
            flash('Yacht added successfully.', 'success')
        elif action == 'edit':
            y = Yacht.query.get_or_404(request.form['yacht_id'])
            y.name = request.form['name'].strip()
            y.model = request.form.get('model', '').strip()
            y.capacity = int(request.form.get('capacity', 10))
            y.length_m = float(request.form.get('length_m', 0) or 0)
            y.description = request.form.get('description', '').strip()
            y.price_per_hour = float(request.form.get('price_per_hour', 0) or 0)
            y.image_url = request.form.get('image_url', '').strip()
            y.is_available = 'is_available' in request.form
            db.session.commit()
            flash('Yacht updated successfully.', 'success')
        elif action == 'toggle':
            y = Yacht.query.get_or_404(request.form['yacht_id'])
            y.is_available = not y.is_available
            db.session.commit()
            flash('Yacht status updated.', 'info')
        return redirect(url_for('admin_yachts'))
    yachts = Yacht.query.order_by(Yacht.name).all()
    return render_template('admin/yachts.html', yachts=yachts)


@app.route('/admin/customers')
@login_required
@admin_required
def admin_customers():
    search = request.args.get('search', '').strip()
    q = db.session.query(
        Booking.customer_name, Booking.customer_email, Booking.customer_phone,
        Booking.client_company,
        db.func.count(Booking.id).label('total'),
        db.func.sum(db.case((Booking.status == 'approved', 1), else_=0)).label('approved'),
        db.func.max(Booking.created_at).label('last_booking'),
    ).group_by(Booking.customer_email)
    if search:
        q = q.filter(db.or_(
            Booking.customer_name.ilike(f'%{search}%'),
            Booking.customer_email.ilike(f'%{search}%'),
            Booking.client_company.ilike(f'%{search}%'),
        ))
    customers = q.order_by(db.func.max(Booking.created_at).desc()).all()
    return render_template('admin/customers.html', customers=customers, search=search)


@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_users():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            uname = request.form['username'].strip()
            if User.query.filter_by(username=uname).first():
                flash('Username already exists.', 'danger')
            else:
                u = User(
                    username=uname,
                    email=request.form['email'].strip().lower(),
                    name=request.form['name'].strip(),
                    department=request.form.get('department', '').strip(),
                    job_title=request.form.get('job_title', '').strip(),
                    role=request.form.get('role', 'staff'),
                )
                u.set_password(request.form['password'])
                db.session.add(u)
                db.session.commit()
                flash('Staff account created successfully.', 'success')
        elif action == 'toggle':
            u = User.query.get_or_404(request.form['user_id'])
            if u.id != current_user.id:
                u.is_active = not u.is_active
                db.session.commit()
                flash('Account status updated.', 'info')
        elif action == 'edit':
            u = User.query.get_or_404(request.form['user_id'])
            u.name = request.form['name'].strip()
            u.email = request.form['email'].strip().lower()
            u.department = request.form.get('department', '').strip()
            u.job_title = request.form.get('job_title', '').strip()
            u.role = request.form.get('role', u.role)
            db.session.commit()
            flash(f'{u.name} has been updated.', 'success')
        elif action == 'delete':
            u = User.query.get_or_404(request.form['user_id'])
            if u.id != current_user.id:
                db.session.delete(u)
                db.session.commit()
                flash('Account deleted.', 'info')
        elif action == 'reset_password':
            u = User.query.get_or_404(request.form['user_id'])
            new_pw = request.form.get('new_password', '').strip()
            if new_pw:
                u.set_password(new_pw)
                db.session.commit()
                flash(f'Password for {u.name} has been reset.', 'success')
        return redirect(url_for('admin_users'))
    users = User.query.order_by(User.role, User.name).all()
    return render_template('admin/users.html', users=users)


# ---------------------------------------------------------------------------
# Email helpers
# ---------------------------------------------------------------------------
def _send_async(msg):
    """Send a Flask-Mail Message in a background thread so the HTTP request
    returns immediately. Errors are logged, never raised."""
    import threading

    def _worker(m):
        # `app` is already the real Flask instance, no proxy unwrap needed
        with app.app_context():
            try:
                mail.send(m)
            except Exception as e:
                app.logger.warning(f'Async email send failed: {e}')

    t = threading.Thread(target=_worker, args=(msg,), daemon=True)
    t.start()


def _send_confirmation(booking):
    if not app.config.get('MAIL_USERNAME'):
        return
    try:
        msg = Message(
            subject=f'[Tomorrow World] Booking Confirmation #{booking.booking_ref}',
            recipients=[booking.customer_email],
        )
        msg.html = render_template('emails/booking_confirmation.html', booking=booking)
        _send_async(msg)
    except Exception as e:
        app.logger.warning(f'Email build failed: {e}')


def _send_status_email(booking):
    if not app.config.get('MAIL_USERNAME'):
        return
    try:
        if booking.status == 'approved':
            status_txt = 'Approved'
        elif booking.status == 'rejected':
            status_txt = 'Not Approved'
        else:
            status_txt = 'Update'
        msg = Message(
            subject=f'[Tomorrow World] Booking {status_txt} #{booking.booking_ref}',
            recipients=[booking.customer_email],
        )
        msg.html = render_template('emails/booking_status.html', booking=booking)
        _send_async(msg)
    except Exception as e:
        app.logger.warning(f'Email build failed: {e}')


def _admin_recipients():
    """Return every email address that should receive admin notifications.

    Combines every active admin user in the DB with the ADMIN_EMAIL config
    value (as a fallback / catch-all). Deduplicated, case-insensitive.
    """
    emails = []
    try:
        admins = User.query.filter_by(role='admin', is_active=True).all()
        emails.extend([u.email for u in admins if u.email])
    except Exception as e:
        app.logger.warning(f'Could not load admin users for notification: {e}')
    fallback = app.config.get('ADMIN_EMAIL')
    if fallback:
        emails.append(fallback)
    # Deduplicate while preserving order, case-insensitive
    seen = set()
    unique = []
    for e in emails:
        key = e.strip().lower()
        if key and key not in seen:
            seen.add(key)
            unique.append(e.strip())
    return unique


def _send_admin_notification(booking):
    """Notify every admin that a new booking is waiting for review."""
    if not app.config.get('MAIL_USERNAME'):
        return
    recipients = _admin_recipients()
    if not recipients:
        return
    try:
        review_url = url_for('admin_booking_detail', bid=booking.id, _external=True)
        msg = Message(
            subject=f'[Tomorrow World] New Booking Pending Review #{booking.booking_ref}',
            recipients=recipients,
        )
        msg.html = render_template(
            'emails/admin_new_booking.html',
            booking=booking,
            review_url=review_url,
        )
        _send_async(msg)
    except Exception as e:
        app.logger.warning(f'Admin notification build failed: {e}')


def _send_cancellation(booking, cancelled_by='admin', reason=''):
    """Send a cancellation email to both the customer and the admin mailbox."""
    if not app.config.get('MAIL_USERNAME'):
        return
    try:
        # Customer copy
        msg_cust = Message(
            subject=f'[Tomorrow World] Booking Cancelled #{booking.booking_ref}',
            recipients=[booking.customer_email],
        )
        msg_cust.html = render_template(
            'emails/booking_cancellation.html',
            booking=booking,
            recipient='customer',
            cancelled_by=cancelled_by,
            cancellation_reason=reason,
        )
        _send_async(msg_cust)

        # Admin copy — send to every admin except if they ARE the customer
        customer_email_lower = (booking.customer_email or '').strip().lower()
        admin_recipients = [
            e for e in _admin_recipients()
            if e.strip().lower() != customer_email_lower
        ]
        if admin_recipients:
            msg_admin = Message(
                subject=f'[Tomorrow World] Booking Cancelled #{booking.booking_ref}',
                recipients=admin_recipients,
            )
            msg_admin.html = render_template(
                'emails/booking_cancellation.html',
                booking=booking,
                recipient='admin',
                cancelled_by=cancelled_by,
                cancellation_reason=reason,
            )
            _send_async(msg_admin)
    except Exception as e:
        app.logger.warning(f'Cancellation email build failed: {e}')


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    with app.app_context():
        init_db()
    port = int(os.environ.get('PORT', 8080))
    app.run(debug=True, host='0.0.0.0', port=port)
